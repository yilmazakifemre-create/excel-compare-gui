import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

def select_file(entry):
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    entry.delete(0, tk.END)
    entry.insert(0, file_path)

def color_and_save(df1, df2, out_path):
    same = []
    diff = []

    for a in df1:
        if a in df2:
            same.append(a)
        else:
            diff.append(a)
    for b in df2:
        if b not in df1:
            diff.append(b)

    # Aynı olanları yan yana yaz
    max_len = max(len(df1), len(df2))
    df_out = pd.DataFrame({
        "Dosya 1": df1 + [""] * (max_len - len(df1)),
        "Dosya 2": df2 + [""] * (max_len - len(df2))
    })
    df_out.to_excel(out_path, index=False)

    # Hücreleri renklendir
    wb = load_workbook(out_path)
    ws = wb.active
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row in range(2, max_len + 2):
        a_val = ws.cell(row=row, column=1).value
        b_val = ws.cell(row=row, column=2).value

        if a_val == b_val and a_val not in ("", None):
            ws.cell(row=row, column=1).fill = green
            ws.cell(row=row, column=2).fill = green
        else:
            if a_val not in ("", None):
                ws.cell(row=row, column=1).fill = red
            if b_val not in ("", None):
                ws.cell(row=row, column=2).fill = red

    wb.save(out_path)


def compare_excels(file1, file2, sheet, col1, col2, start1, end1, start2, end2, search_value):
    try:
        # header=None ekledik → başlık satırı olmasa da çalışır
        df1 = pd.read_excel(file1, sheet_name=sheet, usecols=[col1], header=None)
        df2 = pd.read_excel(file2, sheet_name=sheet, usecols=[col2], header=None)

        list1 = df1.iloc[start1 - 1:end1, 0].dropna().astype(str).tolist()
        list2 = df2.iloc[start2 - 1:end2, 0].dropna().astype(str).tolist()

        if search_value:
            list1 = [x for x in list1 if search_value.lower() in x.lower()]
            list2 = [x for x in list2 if search_value.lower() in x.lower()]

        out_path = "karsilastirma_sonucu.xlsx"
        color_and_save(list1, list2, out_path)
        messagebox.showinfo("Tamamlandı", f"Karşılaştırma tamamlandı!\nSonuç: {out_path}")
    except Exception as e:
        messagebox.showerror("Hata", f"Karşılaştırma sırasında hata oluştu:\n{e}")


def start_compare():
    if not file1_entry.get() or not file2_entry.get():
        messagebox.showwarning("Eksik Bilgi", "Lütfen iki Excel dosyası da seçin.")
        return

    compare_excels(
        file1_entry.get(), file2_entry.get(),
        sheet_entry.get(), col1_entry.get(), col2_entry.get(),
        int(start1_entry.get()), int(end1_entry.get()),
        int(start2_entry.get()), int(end2_entry.get()),
        search_entry.get()
    )


root = tk.Tk()
root.title("Akif Emre Yılmaz – Excel Karşılaştırma")
root.geometry("600x520")

tk.Label(root, text="Excel 1 Seç:").pack()
file1_entry = tk.Entry(root, width=60)
file1_entry.pack()
tk.Button(root, text="Gözat", command=lambda: select_file(file1_entry)).pack()

tk.Label(root, text="Excel 2 Seç:").pack()
file2_entry = tk.Entry(root, width=60)
file2_entry.pack()
tk.Button(root, text="Gözat", command=lambda: select_file(file2_entry)).pack()

tk.Label(root, text="Sayfa Adı (örnek: Sheet1):").pack()
sheet_entry = tk.Entry(root)
sheet_entry.insert(0, "Sheet1")
sheet_entry.pack()

tk.Label(root, text="Sütunlar (örnek: A, B):").pack()
col_frame = tk.Frame(root)
col_frame.pack()
col1_entry = tk.Entry(col_frame, width=5)
col1_entry.insert(0, "A")
col1_entry.pack(side="left", padx=5)
col2_entry = tk.Entry(col_frame, width=5)
col2_entry.insert(0, "A")
col2_entry.pack(side="left", padx=5)

tk.Label(root, text="Satır Aralıkları (örn: 1-100 ve 1-200):").pack()
range_frame = tk.Frame(root)
range_frame.pack()
start1_entry = tk.Entry(range_frame, width=5)
start1_entry.insert(0, "1")
start1_entry.pack(side="left", padx=2)
end1_entry = tk.Entry(range_frame, width=5)
end1_entry.insert(0, "100")
end1_entry.pack(side="left", padx=2)
start2_entry = tk.Entry(range_frame, width=5)
start2_entry.insert(0, "1")
start2_entry.pack(side="left", padx=2)
end2_entry = tk.Entry(range_frame, width=5)
end2_entry.insert(0, "200")
end2_entry.pack(side="left", padx=2)

tk.Label(root, text="Arama (isteğe bağlı):").pack()
search_entry = tk.Entry(root, width=40)
search_entry.pack()

tk.Button(root, text="Karşılaştır", command=start_compare, bg="#4CAF50", fg="white").pack(pady=15)

root.mainloop()
